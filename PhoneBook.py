##############################################################################
# Phone Book
# Allows you to add, subtract, and replace contact information, similar to a 
# contact list in your phone
##############################################################################

from openpyxl import load_workbook #read book
from openpyxl import Workbook #create book
import pandas as pd 

#Create A New Phone Book
def Create_Book():
    wb = Workbook()
    destination_filename = "PhoneBook.xlsx"
    ws1 = wb.active
    ws1.title = "Contact Information" #title the page
    
    ws1["A1"] = "FN" 
    ws1["B1"] = "LN"
    ws1["C1"] = "CN"
    ws1["D1"] = "HN"
    ws1["E1"] = "WN"
    ws1["F1"] = "E"
    ws1["G1"] = "R"
    
    wb.save(filename = destination_filename) #saves the excel sheet
    
    return(wb)
    #add comment here
#Open Existing Phone Book 
def Open_Book():
    print("What is the file name of your current book?")
    destination_filename = input("Answer:") + ".xlsx"
    
    while True:
        try: 
            wb = load_workbook(destination_filename)
            contact_df = Sheet_To_Frame(destination_filename) #making sheet into a pandas data frame
            break
        
        except:
            print("I'm sorry but there was an issue opening the file you specified. Please type it again.")
            destination_filename = input("Answer:")
            continue
        
    return(contact_df)   

#Excel Sheet to DataFrame 
def Sheet_To_Frame(excel_name): 
    contact_list = pd.read_excel(excel_name)
    return(contact_list)
    

def Show_Contacts(contact_list): 
    print(contact_list)
    pass

#Add to the DataFrame
def Add_Contact(contact_df,fn,ln,cell_num,home_num,work_num,email,relationship):
    new_contact_df = contact_df.append([{'FN':fn,'LN':ln,'CN':cell_num,'HN':home_num,'WN':work_num,'E':email,'R':relationship}],ignore_index=True)
    # ^^ append returns a new data frame
    return(new_contact_df)

#Subtract from the DataFrame 
def Remove_Contact():
    pass

#Replace Information from the DataFrame
def Edit_Contact():
    pass

#View All Contacts 
def View_Contacts(): 
    pass

#Search for Contact
def Search_Contact():
    pass
#Changes the name of their phone book
def Change_Book_Name():
    pass 

#Save
def Save_Book():
    pass

#Ask if they are a new user then make a new excel file 
#If they aren't a new user then let them see on an existing file (will have to \ask for the file)

print("Are you a new user? Please answer yes or no.")
contact_df = pd.DataFrame([{'FN':0,'LN':0,'CN':0,'HN':0,'WN':0,'E':0,'R':0}])
not_inside = True 
while True:
    
    
    while True and not_inside: 
        new_user_q = input("Answer[y/n]:")
        new_user_q = new_user_q.lower()
        
        if new_user_q == "yes" or new_user_q == "y": 
            wb = Create_Book()
            print(wb)
            print("Your new contact book has been created")
            print("The books name is 'PhoneBook'")
            print()
            contact_df = Open_Book()
            break
            
        elif new_user_q == "no" or new_user_q == "n":
            contact_df = Open_Book()
            print("Your contact book has been retrieved")
            break 
        
        else: 
            print("Sorry that answer was incorrect. Please try again")
            continue
    not_inside = False    
    print()    
    print("What would you like to do?")
    print("Category Options\n A : Add New Contact \n R : Remove Contact \n E : Edit Existing Contact \n V : View Contacts \n S : Search For Contact \n")
    
    option_answer = input("Choice: ").lower()

    if option_answer == "a": 
        '''add new contact'''
        print("Please provide their information.")
        print()
        fn = input("First Name: ")
        ln = input("Last Name: ")
        cell_num = input("Cell PhoneNumber: ")
        home_num = input("Home phoneNumber: ")
        work_num = input("Work PhoneNumber: ")
        email = input("Email: ")
        relationship = input("Relationship: ")
        print()
        
        contact_df = Add_Contact(contact_df,fn,ln,cell_num,home_num,work_num,email,relationship)
        Show_Contacts(contact_df)
         
    
    elif option_answer == "r": 
        '''remmove contact'''
        pass 
    
    elif option_answer == "e": 
        '''edit existing contact'''
        contact_df.to_excel('PhoneBook.xlsx',sheet_name='Contact Information',index=False)
        # index=false prevents an unamed column from being created
        break
    
    elif option_answer == "v": 
        '''view contacts'''
        Show_Contacts(contact_df) 
        
    elif option_answer == "s": 
        '''search for contact'''
        pass
    else:
        print("Sorry, that is not an option on the menue. Please try again")
        continue

    

        

    # Category Options 
    # A : Add New Contact 
    # R : Remove Contact 
    # E : Edit Existing Contact 
    # V : View Contacts 
    # S : Search For Contact 
